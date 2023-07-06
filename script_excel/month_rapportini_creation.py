import openpyxl
from datetime import datetime
import calendar
import os
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo

# Creazione del file Excel
workbook = openpyxl.Workbook()
sheet = workbook.active

# Ottenere i giorni del mese corrente
now = datetime.now()
year = now.year
month = now.month
num_days = calendar.monthrange(year, month)[1]
month_name = now.strftime("%B")

# Aggiunta delle righe argomento
sheet["A2"] = "argomento 1"
sheet["A3"] = "argomento 2"

# Aggiunta dei giorni come colonne
for day in range(1, num_days + 1):
    column_letter = get_column_letter(day + 1)  # day + 1 per evitare di sovrascrivere la colonna A
    cell = f"{column_letter}1"
    sheet[cell] = f"{day} {month_name}"

# Creazione della tabella
table_range = f"A1:{get_column_letter(num_days + 1)}2"
table = Table(displayName="Table1", ref=table_range)
table.tableStyleInfo = TableStyleInfo(name="TableStyleMedium9", showFirstColumn=False, showLastColumn=False,
                                      showRowStripes=True, showColumnStripes=False)
sheet.add_table(table)

# Imposta larghezza delle colonne
for col in sheet.columns:
    max_length = 0
    column = col[0].column_letter
    for cell in col:
        try:
            if len(str(cell.value)) > max_length:
                max_length = len(cell.value)
        except:
            pass
    adjusted_width = (max_length + 2) * 1.2
    sheet.column_dimensions[column].width = adjusted_width

# Creazione della directory se non esiste
directory = "file_rapportini"
if not os.path.exists(directory):
    os.makedirs(directory)

# Salvataggio del file Excel
filename = f"file_rapportini/tabella_{month_name.lower()}.xlsx"
workbook.save(filename)
print(f"File {filename} creato correttamente.")